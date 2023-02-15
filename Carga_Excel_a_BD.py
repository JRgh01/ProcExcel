from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
import mysql.connector


def funCargaExcel(strNomArc):
#    Archivo = "C:\Temp\Python\CargaDatos\Prueba.xlsx"
    Archivo = "ExcArchivo\\" + strNomArc + ".xlsx"
    print("Cargando Excel de datos: ", Archivo)    
    Libro = Workbook()    
    Libro = load_workbook(Archivo)
    Hoja = Libro.active
    return(Libro, Hoja)

def funCierraExcel(Libro):
    Libro.close()

def funLeeLinea(Hoja, Linea):
    linEx = []
    Columna = 1
    while (Columna < (INTCOLMAX+1)):
        linEx.append(str(Hoja.cell(row = Linea, column = Columna).value))
        if linEx[Columna-1] == "None":
            linEx[Columna-1] = ""
        Columna = Columna + 1
    return(linEx)

def funValRegPerNSSCURP(NSS, CURP):
    #strValPerNSS = "Select ID_Persona, CURP, NSS From Persona Where CURP = "
    curPersona = cnxDB.cursor()
    #Si existe NSS,CURP, obtener el ID_Persona
    strSQL = strValPerNSS + "'" + CURP + "' AND NSS = '" + NSS + "'"
    curPersona.execute(strSQL)
    resSQL = curPersona
    banReg = False
    idxPer = 0
    for valores in resSQL:
        idxPer  = valores[0]
        if CURP == valores[1]:
            banReg = True
            break
    for valores in resSQL:
        aux=0
    curPersona.close()
    if banReg:
        return(idxPer, CURP)
    return(idxPer, "")

def funInsRegPer(Linea):
    RFC = Linea[INTCURP]
    RFC = RFC[:10]
    curPersona = cnxDB.cursor()
    strCad = ""
    for aux in LISINT_PERSONA:
        strCad = strCad + "'" + Linea[aux] + "', "
    strSQL = strINSPer + strCad + "'" + RFC + "')"
    curPersona.execute(strSQL)
    IDPer = 0
    IDPer = curPersona.lastrowid
#    print(strSQL, IDPer)
    curPersona.close()
    return(IDPer)

def funInsRegCorreo(IdPer, Correo, Prov):
#    strINSCorr = "INSERT INTO Correo (ID_Persona, Correo, Procedencia) VALUES ("
    curPersona = cnxDB.cursor()
    strSQL = strINSCorr + str(IdPer) + ", '" + Correo + "', " + str(Prov) + ")"
    curPersona.execute(strSQL)
    curPersona.close()

def funInsRegDom(IdPer, Calle, Colonia, DomMun, DomEdo):
    #strINSDom = "INSERT INTO Domicilio (ID_Persona, DomCalle, DomColonia, DomMun, DomEdo) VALUES ("
    curPersona = cnxDB.cursor()
    if Calle != "" or Colonia != "" or DomMun != "" or DomEdo != "":
        strSQL = strINSDom + str(IdPer) + ", '" + Calle + "', '" + Colonia + "', '" + DomMun + "', '" + DomEdo + "')"
        curPersona.execute(strSQL)
        curPersona.close()

def funInsRegExcArc(Archivo):
    curPersona = cnxDB.cursor()
    
    strSQL = "Select ID_Excel, NomArchivo From ExcArchivo Where NomArchivo = '" + Archivo + "'"
    curPersona.execute(strSQL)
    resSQL = curPersona
    idxErr = 0
    for valores in resSQL:
        idxErr  = valores[0]
        break

    if idxErr == 0:
        strINS = "INSERT INTO ExcArchivo (NomArchivo) VALUES ("
        strSQL = strINS + "'" + Archivo + "')"
        curPersona.execute(strSQL)
        idxErr = 0
        idxErr = curPersona.lastrowid
   
    curPersona.close()
    return idxErr

def funInsRegExcPer(IdPer, IdExc):
    curPersona = cnxDB.cursor()

    strSQL = "Select ID_Persona, ID_Excel From ExcPersona "
    strSQL = strSQL + "Where ID_Persona = " + str(IdPer) + " AND ID_Excel = " + str(IdExc)
    curPersona.execute(strSQL)
    resSQL = curPersona
    idxErr = 0
    for valores in resSQL:
        idxErr  = valores[0]
        break

    if idxErr == 0:
        strINS = "INSERT INTO ExcPersona (ID_Persona, ID_Excel) VALUES ("
        strSQL = strINS + str(IdPer) + ", " + str(IdExc) + ")"
        curPersona.execute(strSQL)
        idxErr = 0
        idxErr = curPersona.lastrowid
   
    curPersona.close()

def funInsRegTel(IdPer, Telefono, Prov):
#    strINSTel = "INSERT INTO Telefono (ID_Persona, Telefono, Procedencia) VALUES ("
    curPersona = cnxDB.cursor()
    strSQL = strINSTel + str(IdPer) + ", '" + Telefono + "', " + str(Prov) + ")"
    curPersona.execute(strSQL)
    curPersona.close()

def funInsRegCapP(IdPer, Comentario, Monto, PreVig):
    LisPreVig = []
    intCom = 5
    if Comentario.find("social") > -1:
        intCom = 1
    elif Comentario.find("CURP") > -1:
        intCom = 2
    elif Comentario.find("servicio") > -1:
        intCom = 3
    else:
        intCom = 4
    if PreVig != "":
        intVig = PreVig.find(" | ")        
        while (intVig > -1): #0123 | 012 | 
            LisPreVig.append(PreVig[0:intVig])
            PreVig = PreVig[intVig + 3:]
            intVig = PreVig.find(" | ")
        for strAux in LisPreVig:
            curPersona = cnxDB.cursor()
            strSQL = strINSCapP + str(IdPer) + ", '" + strAux + "', '" + Monto + "', " + str(intCom) + ")"
            curPersona.execute(strSQL)
            curPersona.close()
    else:        
#    strINSCapP = "INSERT INTO CapacidadPago (ID_Persona, PrestamoVig, Monto, ID_Comentario) VALUES ("
        curPersona = cnxDB.cursor()
        strSQL = strINSCapP + str(IdPer) + ", '" + PreVig + "', '" + Monto + "', " + str(intCom) + ")"
        curPersona.execute(strSQL)
        curPersona.close()

def ProgPrincipal():
    try:
        intCuentaInsPer = 0
        intLin = 2
        strNomArc = ""
        while strNomArc == "":
            strNomArc = input('Nombre de Archivo de Excel a procesar: ')

        xlsLibro, xlsHoja = funCargaExcel(strNomArc)
        if xlsHoja == None:
            return
        idxExc = funInsRegExcArc(strNomArc)
        print("Inicia Proceso ...")
        xlsLinea = funLeeLinea(xlsHoja, intLin)
        while (xlsLinea[INTNSS] != "" or xlsLinea[INTCURP] != ""):
            if (intLin % 10000) == 0:
                print("LLevo: ", intLin)

            idxPer, auxCURP = funValRegPerNSSCURP(xlsLinea[INTNSS], xlsLinea[INTCURP])
            if idxPer == 0:
                idxPer = funInsRegPer(xlsLinea)
                intCuentaInsPer = intCuentaInsPer + 1
            elif idxPer > 0:
                if xlsLinea[INTCURP] != auxCURP:
                    idxPer = funInsRegPer(xlsLinea)
                    intCuentaInsPer = intCuentaInsPer + 1

            if idxPer > 0:
                funInsRegDom(idxPer,xlsLinea[INTDomCalle],xlsLinea[INTDomCol], xlsLinea[INTDomMun], xlsLinea[INTDomEdo] )
                funInsRegExcPer(idxPer, idxExc)

            intLin = intLin + 1
            if intLin > LINMAX:
                break
            xlsLinea = funLeeLinea(xlsHoja, intLin)
        
        print("Registros leídos del EXCEL  : ", intLin-2)
        print("Registros Insertados Persona: ", intCuentaInsPer)
        funCierraExcel(xlsLibro)
    except Exception as error:
        print(error)

try:
#strValPer = "Select count(*) From Persona Where NSS = "        
#Si existe NSS,CURP, obtener el ID_Persona


#        strSQL = strINSPer + "'" + str(strNSS) + "', '" + strCURP + "')"
#        print(strSQL)
#        curPersona.execute(strSQL)
#        idPersona = curPersona.lastrowid
#        print(idPersona)

#LINMAX = 2
    LINMAX = 200000
    INTCOLMAX      = 9    
    INTCURP        = 0
    INTNSS         = 1
    INTPaterno     = 2
    INTMaterno     = 3
    INTNombre      = 4
    INTDomCalle    = 5
    INTDomCol      = 6
    INTDomMun      = 7
    INTDomEdo      = 8


    LISINT_NOCOL   = [0,3,6,7,8,9,13,15,21,22]
    LISINT_PERSONA   = [1,0,2,3,4]    

    strValPerNSS = "Select ID_Persona, CURP From Persona Where CURP = "
    strINSPer = "INSERT INTO Persona (NSS, CURP, ApPaterno, ApMaterno, Nombre, RFC ) VALUES ("
    strINSDom = "INSERT INTO Domicilio (ID_Persona, DomCalle, DomColonia, DomMun, DomEdo) VALUES ("

    strINSCorr = "INSERT INTO Correo (ID_Persona, Correo, Procedencia) VALUES ("    
    strINSTel = "INSERT INTO Telefono (ID_Persona, Telefono, Procedencia) VALUES ("
    strINSCapP = "INSERT INTO CapacidadPago (ID_Persona, PrestamoVig, Monto, ID_Comentario) VALUES ("    

    cnxDB = mysql.connector.connect(user='root', password='',
                                    host='127.0.0.1',
                                    database='SisPre')

    ProgPrincipal()
#    xlsLibro.save(strArc)

# Make sure data is committed to the database
    cnxDB.commit()

    cnxDB.close()
    print("Termina Proceso ...")
except Exception as error:
    print(error)
