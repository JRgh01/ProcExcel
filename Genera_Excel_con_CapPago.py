from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import mysql.connector


def funCargaExcel(nomArc):
#    Archivo = "C:\Temp\Python\CargaDatos\Prueba.xlsx"
#    Archivo = ".\ConsultaGral.xlsx"
    Archivo = "ExcSalida\\" + nomArc
    print("Generando Excel de datos: ", Archivo)    
    Libro = Workbook()    
    Libro = load_workbook(Archivo)
    Hoja = Libro.active
    return(Libro, Hoja)

def funCierraExcel(Libro, Archivo):
    strArcSal = "ExcSalida\\" + "C_Sal.xlsx"
    Libro.save(strArcSal)
    Libro.close()

def funLeeLinea(Hoja, Linea):
    linEx = []
    Columna = 1
    while (Columna < (INTCOLMAX+1)):
        linEx.append(str(Hoja.cell(row = Linea, column = Columna).value))
        if linEx[Columna-1] == "None":
            linEx[Columna-1] = ""
        Columna = Columna + 1
    return(linEx)

def funLeeTablaReg(intTabla, IdPer):
    dbCursor = cnxDB.cursor()  
    if intTabla == 0:
        strSQL = "Select ID_Persona, NSS, CURP, ApPaterno, ApMaterno, Nombre, ID_Comentario From Persona Where ID_Persona = " + str(IdPer)
    elif intTabla == 1:
        strSQL = "Select DomCalle, DomColonia, DomCP, DomMun, DomEdo From Domicilio Where ID_Persona = " + str(IdPer)
    elif intTabla == 2:
        strSQL = "Select Correo From Correo Where ID_Persona = " + str(IdPer)
    elif intTabla == 3:
        strSQL = "Select Telefono From Telefono Where ID_Persona = " + str(IdPer)
    elif intTabla == 4:
        strSQL = "Select Monto, year(fecha), month(fecha), day(fecha) From CapacidadPago Where ID_Persona = " + str(IdPer)
    elif intTabla == 5:
        strSQL = "Select PrestamoVig From PrestamoVig Where ID_Persona = " + str(IdPer)
    elif intTabla == 6:
        strSQL = "Select Comentario From Comentario Where ID_Comentario = " + str(IdPer)

    dbCursor.execute(strSQL)
    auxCur = dbCursor
    
    intReg = 0

    lisCampo = []    
    for resultado in auxCur:
        for Campo in resultado:
            lisCampo.append(Campo)
        intReg = intReg + 1

    dbCursor.close()
    return (intReg, lisCampo)

def funInsRegCapP(IdPer, Comentario, Monto, PreVig):
    LisPreVig = []
    intCom = 5
    if Comentario.find("social") > -1:
        intCom = 1
    elif Comentario.find("CURP") > -1:
        intCom = 2
    elif Comentario.find("servicio") > -1:
        intCom = 3
    else:
        intCom = 4
    if PreVig != "":
        intVig = PreVig.find(" | ")        
        while (intVig > -1): #0123 | 012 | 
            LisPreVig.append(PreVig[0:intVig])
            PreVig = PreVig[intVig + 3:]
            intVig = PreVig.find(" | ")
        for strAux in LisPreVig:
            curPersona = cnxDB.cursor()
            strSQL = strINSCapP + str(IdPer) + ", '" + strAux + "', '" + Monto + "', " + str(intCom) + ")"
            curPersona.execute(strSQL)
            curPersona.close()

def ProgPrincipal():
    try:
        intCuentaInsPer = 0
        intLin = 2
        #strArc = ""
        #while strArc == "":
        #    strArc = input('Nombre de Archivo de Excel a procesar: ')
        strArc = "Consulta.xlsx"
        xlsLibro, xlsHoja = funCargaExcel(strArc)
        print("Inicia Proceso ...")

        dbPersona = cnxDB.cursor()  
        strSQL = "Select Persona.ID_Persona, NSS, CURP, ApPaterno, ApMaterno, Nombre, ID_Comentario "
        strSQL = strSQL + "From Persona INNER JOIN ExcPersona "
        strSQL = strSQL + "  ON Persona.ID_Persona = ExcPersona.ID_Persona "
        strSQL = strSQL + " Where ExcPersona.ID_Excel = 2"

        #strSQL = "Select ID_Persona, NSS, CURP, ApPaterno, ApMaterno, Nombre, ID_Comentario From Persona Where ID_Persona = 3"
        dbPersona.execute(strSQL)
        auxCur = dbPersona
    
        IDPersona = []
        for resultado in auxCur:
            IDPersona.append(resultado[0])
        #print(IDPersona)
        dbPersona.close()

        lisPersona = []
        lisDomicilio = []
        lisCorreo = []
        lisTelefono = []
        lisCapPAgo = []
        lisPrestamo = []
        lisComentario = []

        for idPer in IDPersona:
            lisExcel = []
            intRegTab, lisPersona = funLeeTablaReg(0, idPer)    #Persona
            intAux = 0
            while intAux < 6:
                lisExcel.append(lisPersona[intAux])
                intAux = intAux + 1

            intRegTab, lisDomicilio = funLeeTablaReg(1, idPer)  #Domicilio            
            if intRegTab > 0:
                intAux = 1
                while intAux < 6:
                    lisExcel.append(lisDomicilio[intAux-1])
                    intAux = intAux + 1
            else:
                intAux = 1
                while intAux < 6:
                    lisExcel.append("")
                    intAux = intAux + 1

            intRegTab, lisCorreo = funLeeTablaReg(2, idPer)    #Correo
            if intRegTab > 0:
                lisExcel.append(lisCorreo[0])
            else:
                lisExcel.append("")
            
            intRegTab, lisTelefono = funLeeTablaReg(3, idPer)  #Telefono
            if intRegTab == 0:
                lisExcel.append("")
                lisExcel.append("")                                
            elif intRegTab == 1:
                lisExcel.append(lisTelefono[0])
                lisExcel.append("")                                
            else:
                lisExcel.append(lisTelefono[0])
                lisExcel.append(lisTelefono[1])                

            intRegTab, lisCapPAgo = funLeeTablaReg(4, idPer)  #CapacidadPago
            if intRegTab > 0:
                lisExcel.append(lisCapPAgo[0])
            else:
                lisExcel.append("")

            intRegTab, lisPrestamo = funLeeTablaReg(5, idPer)  #Prestamo
            if intRegTab > 0:
                intAux = 1
                strPre = ""
                strPre = strPre + lisPrestamo[0]
                while intAux < intRegTab:
                    strPre = strPre + " | " + lisPrestamo[intAux]
                    intAux = intAux + 1

                lisExcel.append(strPre)
            else:
                lisExcel.append("")

            if lisPersona[6] > 0:
                intRegTab, lisComentario = funLeeTablaReg(6, lisPersona[6])  #Comentario
                if intRegTab > 0:
                    lisExcel.append(lisComentario[0])
                else:
                    lisExcel.append("")

            #print(lisExcel)
            intCol = 0
            for auxEx in lisExcel:
                xlsHoja.cell(row = intLin, column = intCol+1).value = auxEx
                intCol = intCol + 1
            intLin = intLin + 1

        print("Registros leídos del EXCEL  : ", intLin-2)
        print("Registros Insertados Persona: ", intCuentaInsPer)

        funCierraExcel(xlsLibro, strArc)

    except Exception as error:
        print(error)

try:
#strValPer = "Select count(*) From Persona Where NSS = "        
#Si existe NSS,CURP, obtener el ID_Persona


#        strSQL = strINSPer + "'" + str(strNSS) + "', '" + strCURP + "')"
#        print(strSQL)
#        curPersona.execute(strSQL)
#        idPersona = curPersona.lastrowid
#        print(idPersona)

#LINMAX = 2
    LINMAX = 200000
    INTCOLMAX      = 9    
    INTCURP        = 0
    INTNSS         = 1
    INTPaterno     = 2
    INTMaterno     = 3
    INTNombre      = 4
    INTDomCalle    = 5
    INTDomCol      = 6
    INTDomMun      = 7
    INTDomEdo      = 8


    LISINT_NOCOL   = [0,3,6,7,8,9,13,15,21,22]
    LISINT_PERSONA   = [1,0,2,3,4]    

    strValPerNSS = "Select ID_Persona, CURP From Persona Where NSS = "
    strINSPer = "INSERT INTO Persona (NSS, CURP, ApPaterno, ApMaterno, Nombre, RFC ) VALUES ("
    strINSDom = "INSERT INTO Domicilio (ID_Persona, DomCalle, DomColonia, DomMun, DomEdo) VALUES ("

    strINSCorr = "INSERT INTO Correo (ID_Persona, Correo, Procedencia) VALUES ("    
    strINSTel = "INSERT INTO Telefono (ID_Persona, Telefono, Procedencia) VALUES ("
    strINSCapP = "INSERT INTO CapacidadPago (ID_Persona, PrestamoVig, Monto, ID_Comentario) VALUES ("    

    cnxDB = mysql.connector.connect(user='root', password='',
                                    host='127.0.0.1',
                                    database='SiPre')

    ProgPrincipal()


# Make sure data is committed to the database
    cnxDB.commit()

    cnxDB.close()
    print("Termina Proceso ...")
except Exception as error:
    print(error)
